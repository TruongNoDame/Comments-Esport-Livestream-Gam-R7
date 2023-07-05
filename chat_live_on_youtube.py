# Import package
from chat_downloader import ChatDownloader
from openpyxl import Workbook  # Now prepare a xlsx to write crawl data

wb = Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = 'Content'
ws['A2'] = 'Final'
# Now copy and paste something like
url = 'https://www.youtube.com/watch?v=KBEMaiOV8Is&t=3s'  # Only the football live stream which has chat replay
chat = ChatDownloader().get_chat(url=url, start_time="00:19:00", end_time="02:00:00")  # Get chats from url
d = 0
lst = []
# Create a generator
for message in chat:
    """Crawled chat should be like this
    {
    ...
    "message_id": "xxxxxxxxxx",
    "message": "actual message goes here",
    "message_type": "text_message",
    "timestamp": 1613761152565924,
    "time_in_seconds": 1234.56,
    "time_text": "20:34",
    "author": {
        "id": "UCxxxxxxxxxxxxxxxxxxxxxxx",
        "name": "username_of_sender",
        "images": [
            ...
        ],
        "badges": [
            ...
        ]
    },
    ...
    }"""
    # In here we just get 7000 messages
    # if(d>7000):break
    if (len(message['message']) > 1):  # iterate over messages
        string = message['message']  # Just take message content
        import re

        a = re.findall(':.*?:', string)  # We will remove all emote, which start and end with char ':'
        for sub in a:
            string = string.replace(sub, '')
        # Now we try to remove all the same message content by create a substring without space char and check if it existed, write to file
        st = string.replace(' ', '')
        if (string != '' and lst.count(st) == 0):
            ws.append([string])
            print(string)
            lst.append(st)
            d = d + 1
wb.save("/DataCrawl.xlsx")
